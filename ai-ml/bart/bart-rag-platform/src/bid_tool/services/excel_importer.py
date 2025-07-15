"""
Excel importer for BART 3.20 workbook.
Extracts all data, formulas, and business logic while preserving structure.
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import json
import logging
from typing import Dict, Any, List, Tuple, Optional
from pathlib import Path
import re

logger = logging.getLogger(__name__)


class BARTExcelImporter:
    """Import and convert BART 3.20 Excel workbook"""
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.workbook = None
        self.data = {}
        self.formulas = {}
        self.lookups = {}
        self.validations = {}
        
    def import_all(self) -> Dict[str, Any]:
        """Import entire Excel workbook preserving all information"""
        logger.info(f"Starting import of {self.excel_path}")
        
        # Load workbook
        self.workbook = openpyxl.load_workbook(
            self.excel_path, 
            data_only=False,  # Keep formulas
            keep_vba=True
        )
        
        # Process each sheet
        for sheet_name in self.workbook.sheetnames:
            logger.info(f"Processing sheet: {sheet_name}")
            sheet = self.workbook[sheet_name]
            
            # Skip if sheet is hidden (but still extract formulas)
            if sheet.sheet_state == 'hidden':
                self._extract_hidden_sheet_logic(sheet_name, sheet)
            else:
                self._extract_visible_sheet_data(sheet_name, sheet)
                
        # Extract all VLOOKUP references
        self._extract_lookup_tables()
        
        # Extract data validations (dropdowns)
        self._extract_validations()
        
        # Build formula dependency graph
        formula_graph = self._build_formula_dependencies()
        
        return {
            'data': self.data,
            'formulas': self.formulas,
            'lookups': self.lookups,
            'validations': self.validations,
            'formula_dependencies': formula_graph,
            'metadata': {
                'source_file': str(self.excel_path),
                'sheet_count': len(self.workbook.sheetnames),
                'formula_count': len(self.formulas),
                'lookup_count': len(self.lookups)
            }
        }
        
    def _extract_visible_sheet_data(self, sheet_name: str, sheet):
        """Extract data from visible sheets (measurement forms)"""
        
        # Identify sheet type
        sheet_type = self._identify_sheet_type(sheet_name)
        
        # Extract form structure
        form_data = {
            'sheet_name': sheet_name,
            'sheet_type': sheet_type,
            'fields': [],
            'merged_cells': [],
            'formulas': {},
            'dropdowns': []
        }
        
        # Process merged cells (form labels)
        for merged_range in sheet.merged_cells.ranges:
            form_data['merged_cells'].append({
                'range': str(merged_range),
                'value': sheet[merged_range.min_row][merged_range.min_col - 1].value
            })
            
        # Extract all cells with data or formulas
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None or cell.formula is not None:
                    cell_data = {
                        'address': cell.coordinate,
                        'value': cell.value,
                        'formula': cell.formula,
                        'data_type': cell.data_type,
                        'number_format': cell.number_format
                    }
                    
                    # Check if this is a form field
                    if self._is_input_field(cell):
                        form_data['fields'].append({
                            'name': self._get_field_name(sheet, cell),
                            'address': cell.coordinate,
                            'type': self._get_field_type(cell),
                            'required': self._is_required_field(sheet, cell)
                        })
                        
                    # Store formula
                    if cell.formula:
                        formula_key = f"{sheet_name}!{cell.coordinate}"
                        self.formulas[formula_key] = {
                            'formula': cell.formula,
                            'sheet': sheet_name,
                            'cell': cell.coordinate,
                            'dependencies': self._extract_formula_deps(cell.formula)
                        }
                        
        self.data[sheet_name] = form_data
        
    def _extract_hidden_sheet_logic(self, sheet_name: str, sheet):
        """Extract formulas and lookups from hidden sheets"""
        
        hidden_data = {
            'sheet_name': sheet_name,
            'sheet_type': 'calculation',
            'is_hidden': True,
            'formulas': {},
            'lookup_tables': {}
        }
        
        # Identify if this is a pricing table
        if 'pricing' in sheet_name.lower() or 'formula' in sheet_name.lower():
            # Extract as lookup table
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)
            
            # Find table boundaries
            table_data = []
            headers = []
            
            for idx, row in df.iterrows():
                if not row.isna().all():
                    if not headers:
                        headers = [str(h) for h in row if pd.notna(h)]
                    else:
                        table_data.append([v for v in row if pd.notna(v)])
                        
            if headers and table_data:
                lookup_key = self._generate_lookup_key(sheet_name)
                self.lookups[lookup_key] = {
                    'sheet': sheet_name,
                    'headers': headers,
                    'data': table_data,
                    'composite_key': self._detect_composite_key(headers)
                }
                
        # Extract all formulas
        for row in sheet.iter_rows():
            for cell in row:
                if cell.formula:
                    formula_key = f"{sheet_name}!{cell.coordinate}"
                    self.formulas[formula_key] = {
                        'formula': cell.formula,
                        'sheet': sheet_name,
                        'cell': cell.coordinate,
                        'is_hidden': True,
                        'complexity': self._calculate_formula_complexity(cell.formula)
                    }
                    
        self.data[sheet_name] = hidden_data
        
    def _extract_lookup_tables(self):
        """Extract all VLOOKUP table references"""
        
        # Parse all formulas to find VLOOKUP references
        vlookup_pattern = r'VLOOKUP\s*\([^,]+,\s*([^,]+),\s*(\d+)'
        
        for formula_key, formula_data in self.formulas.items():
            formula = formula_data['formula']
            
            # Find VLOOKUP calls
            matches = re.findall(vlookup_pattern, formula, re.IGNORECASE)
            for table_ref, col_index in matches:
                # Clean table reference
                table_ref = table_ref.strip()
                
                # Track usage
                if table_ref not in self.lookups:
                    self.lookups[table_ref] = {
                        'references': [],
                        'columns_used': set()
                    }
                    
                self.lookups[table_ref]['references'].append(formula_key)
                self.lookups[table_ref]['columns_used'].add(int(col_index))
                
    def _extract_validations(self):
        """Extract data validations (dropdowns)"""
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            # Check for data validations
            if hasattr(sheet, 'data_validations'):
                for validation in sheet.data_validations.dataValidation:
                    self.validations[f"{sheet_name}_{validation.sqref}"] = {
                        'sheet': sheet_name,
                        'range': str(validation.sqref),
                        'type': validation.type,
                        'formula': validation.formula1,
                        'allow_blank': validation.allowBlank,
                        'show_dropdown': validation.showDropDown
                    }
                    
    def _build_formula_dependencies(self) -> Dict[str, List[str]]:
        """Build dependency graph for formulas"""
        
        dependencies = {}
        
        for formula_key, formula_data in self.formulas.items():
            deps = []
            formula = formula_data['formula']
            
            # Extract cell references
            cell_refs = re.findall(r'([A-Z]+[0-9]+)', formula)
            sheet_refs = re.findall(r"'?([^'!]+)'?!([A-Z]+[0-9]+)", formula)
            
            # Add dependencies
            for ref in cell_refs:
                deps.append(f"{formula_data['sheet']}!{ref}")
                
            for sheet, cell in sheet_refs:
                deps.append(f"{sheet}!{cell}")
                
            dependencies[formula_key] = list(set(deps))
            
        return dependencies
        
    def _identify_sheet_type(self, sheet_name: str) -> str:
        """Identify the type of sheet based on name"""
        
        name_lower = sheet_name.lower()
        
        if 'measure' in name_lower:
            if 'ext' in name_lower:
                return 'exterior_measurement'
            elif 'int' in name_lower:
                return 'interior_measurement'
            elif 'cabinet' in name_lower:
                return 'cabinet_measurement'
            elif 'holiday' in name_lower:
                return 'holiday_measurement'
            else:
                return 'measurement'
        elif 'formula' in name_lower or 'pricing' in name_lower:
            return 'calculation'
        elif 'crew' in name_lower:
            return 'crew_assignment'
        elif 'checklist' in name_lower:
            return 'checklist'
        elif 'client' in name_lower:
            return 'client_info'
        elif 'how to' in name_lower:
            return 'instructions'
        else:
            return 'other'
            
    def _is_input_field(self, cell) -> bool:
        """Determine if a cell is an input field"""
        
        # Input fields typically have no formula and specific formatting
        return (
            cell.formula is None and
            cell.value is None and
            cell.fill.start_color.index in ['FFFFFF', 'FFFF00', None]  # White/yellow
        )
        
    def _get_field_name(self, sheet, cell) -> str:
        """Extract field name from nearby label"""
        
        # Look for label in cells to the left
        row = cell.row
        for col in range(cell.column - 1, 0, -1):
            label_cell = sheet.cell(row=row, column=col)
            if label_cell.value:
                return str(label_cell.value)
                
        # Look above
        col = cell.column
        for row in range(cell.row - 1, 0, -1):
            label_cell = sheet.cell(row=row, column=col)
            if label_cell.value:
                return str(label_cell.value)
                
        return f"Field_{cell.coordinate}"
        
    def _get_field_type(self, cell) -> str:
        """Determine field type from formatting"""
        
        if cell.number_format:
            if '$' in cell.number_format:
                return 'currency'
            elif '%' in cell.number_format:
                return 'percentage'
            elif '0' in cell.number_format:
                return 'number'
                
        return 'text'
        
    def _is_required_field(self, sheet, cell) -> bool:
        """Check if field is required based on conditional formatting or notes"""
        
        # Check for validation
        if hasattr(cell, 'data_validation') and cell.data_validation:
            return not cell.data_validation.allowBlank
            
        return False
        
    def _calculate_formula_complexity(self, formula: str) -> int:
        """Calculate complexity score for formula"""
        
        score = 0
        
        # Count function calls
        functions = re.findall(r'[A-Z]+\(', formula)
        score += len(functions) * 2
        
        # Count nested parentheses
        max_depth = 0
        current_depth = 0
        for char in formula:
            if char == '(':
                current_depth += 1
                max_depth = max(max_depth, current_depth)
            elif char == ')':
                current_depth -= 1
        score += max_depth * 3
        
        # Count conditions
        score += formula.count('IF') * 5
        
        return score
        
    def _detect_composite_key(self, headers: List[str]) -> Optional[str]:
        """Detect if table uses composite keys"""
        
        # Common patterns for composite keys
        if len(headers) > 2 and headers[0] in ['Type', 'Category', 'Item']:
            return f"{{{headers[0]}}} {{{headers[1]}}}"
            
        return None
        
    def _generate_lookup_key(self, sheet_name: str) -> str:
        """Generate unique key for lookup table"""
        
        # Remove common suffixes
        key = sheet_name.replace(' Sheet', '').replace(' Table', '')
        key = key.replace(' ', '_').lower()
        
        return key